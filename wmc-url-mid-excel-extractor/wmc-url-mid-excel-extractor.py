"""
Wikimedia Commons URL M-ID Excel Extractor
======================================

Reads a Commons file URL column from an Excel sheet, looks up the corresponding
MediaInfo entity IDs (“M-IDs”), and writes the results back **into the same
workbook** by adding/updating two columns:

- ``FileMid``: the MediaInfo identifier in the form ``M{pageid}``
- ``FileMidURL``: the human-readable entity page URL,
  e.g. ``https://commons.wikimedia.org/wiki/Special:EntityPage/M12345``

The script preserves all other sheets in the workbook and replaces only the
target sheet’s contents.

Features
--------
- Robust URL parsing for multiple Commons URL shapes:
  ``/wiki/File:…``, ``?title=File:…``, ``/wiki/Special:FilePath/…``,
  ``/wiki/Special:Redirect/file/…``.
- Batched API requests (≤ 50 titles/request) with retries and exponential backoff.
- Redirect and normalization handling so titles resolve to the correct page.
- One output per input row; unresolved lookups yield ``NOT FOUND`` in both columns.
- Errors and failed lookups are logged to a CSV file.

Inputs
------
- Excel workbook at ``XLSX_PATH``
- Sheet name ``SHEET_NAME`` containing the column ``URL_COLUMN`` with Commons file URLs

Outputs
-------
- The same Excel workbook updated in place with two columns inserted immediately
  after ``URL_COLUMN``:
  - ``MID_COLUMN`` (default: ``FileMid``)
  - ``MID_URL_COLUMN`` (default: ``FileMidURL``)
- A CSV file (``ERRORS_CSV``) with two columns: ``URL_COLUMN`` and ``Error``

Configuration
-------------
Edit the constants at the top of the file:
``XLSX_PATH``, ``SHEET_NAME``, ``URL_COLUMN``, ``MID_COLUMN``, ``MID_URL_COLUMN``,
``ERRORS_CSV``, and ``BATCH_SIZE``. Also set a contact email in ``USER_AGENT`` to
comply with Wikimedia API etiquette.

Usage
-----
1. Close the Excel workbook (it cannot be open while writing).
2. Run the script: ``python mid_extractor.py``
3. Reopen the workbook to see ``FileMid`` and ``FileMidURL`` added next to ``FileURL``.

Requirements
------------
Python 3.9+ with: ``pandas``, ``openpyxl``, ``requests``, ``urllib3``.

Notes
-----
- The M-ID is derived from the MediaWiki pageid (``prop=info``) for the file page.
- ``FileMidURL`` points to the human-readable entity page. If you prefer the
  machine-readable endpoint, use ``Special:EntityData/{mid}.json`` instead.

-----
Author: ChatGPT, prompted by Olaf Janssen, KB (Koninklijke Bibliotheek)
Latest update: 2025-10-17
License: CC0
"""

from __future__ import annotations
import csv
import re
from typing import Dict, Iterable, List, Optional, Tuple
from urllib.parse import urlparse, parse_qs, unquote
import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ===== Configuration (edit these) ============================================
# Inputs
XLSX_PATH = "testfile.xlsx"       # same file used for reading and writing
SHEET_NAME = "FileURLs"            # sheet with the FileURL column
URL_COLUMN = "FileURL"             # input column

# Outputs
# XLSX_PATH is the same for the output
MID_COLUMN = "FileMid"             # output column 1
MID_URL_COLUMN = "FileMidURL"      # output column 2

# Errors
ERRORS_CSV = "errors.csv"
BATCH_SIZE = 50  # MediaWiki allows up to 50 titles/request for non-bots
# ============================================================================

COMMONS_API = "https://commons.wikimedia.org/w/api.php"
ENTITY_PAGE_BASE = "https://commons.wikimedia.org/wiki/Special:EntityPage/"
USER_AGENT = "WikiCommons-MID-Extractor/1.0 (contact: KB, national library of the Netherlands - olaf.janssen@kb.nl)"

def build_session() -> requests.Session:
    """Create a requests session with retry/backoff and a polite User-Agent."""
    session = requests.Session()
    retry = Retry(
        total=6,
        connect=3,
        read=3,
        status=4,
        backoff_factor=0.6,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        raise_on_status=False,
    )
    adapter = HTTPAdapter(max_retries=retry)
    session.mount("https://", adapter)
    session.headers.update({"User-Agent": USER_AGENT})
    return session


FILE_PREFIX_RE = re.compile(r"^file:", re.IGNORECASE)

def extract_title_from_url(url: str) -> Optional[str]:
    """
    Extract 'File:Title.ext' from Commons URL shapes:
      - /wiki/File:Example.jpg
      - /w/index.php?title=File:Example.jpg
      - /wiki/Special:FilePath/Example.jpg
      - /wiki/Special:Redirect/file/Example.jpg
    """
    try:
        if not isinstance(url, str) or not url.strip():
            return None

        p = urlparse(url)
        path = unquote(p.path or "")
        query = parse_qs(p.query or "")

        # 1) /wiki/File:Title
        if "/wiki/" in path and "File:" in path:
            idx = path.lower().find("/wiki/")
            tail = path[idx + len("/wiki/") :]
            if tail.lower().startswith("file:"):
                title = tail.split("/", 1)[0]
                return "File:" + title[5:]

        # 2) ?title=File:Title
        title_param = query.get("title", [None])[0]
        if title_param and FILE_PREFIX_RE.match(title_param):
            return "File:" + title_param[5:]

        # 3) Special helpers
        if "/Special:FilePath/" in path:
            name = path.split("/Special:FilePath/", 1)[1].split("/", 1)[0].strip()
            if name:
                return f"File:{name}"
        if "/Special:Redirect/file/" in path:
            name = path.split("/Special:Redirect/file/", 1)[1].split("/", 1)[0].strip()
            if name:
                return f"File:{name}"

        return None
    except Exception:
        return None


def chunked(seq: Iterable[str], size: int) -> Iterable[List[str]]:
    """Yield lists of length <= size from seq."""
    buf: List[str] = []
    for x in seq:
        buf.append(x)
        if len(buf) >= size:
            yield buf
            buf = []
    if buf:
        yield buf


def fetch_mids_for_titles(
    session: requests.Session, input_titles: List[str], batch_size: int
) -> Tuple[Dict[str, Optional[str]], List[Tuple[str, str]]]:
    """
    Return:
      - mapping {input_title -> "M12345" or None}
      - error log list of (input_title, message)
    """
    results: Dict[str, Optional[str]] = {t: None for t in input_titles}
    errors: List[Tuple[str, str]] = []

    unique_titles = list(dict.fromkeys(input_titles))  # preserve order
    for group in chunked(unique_titles, batch_size):
        try:
            r = session.get(
                COMMONS_API,
                params={
                    "action": "query",
                    "format": "json",
                    "formatversion": "2",
                    "prop": "info",
                    "redirects": "1",
                    "titles": "|".join(group),
                },
                timeout=15,
            )
            r.raise_for_status()
            data = r.json()
        except Exception as e:
            msg = f"HTTP/parse error: {e}"
            for t in group:
                results[t] = None
                errors.append((t, msg))
            continue

        alias: Dict[str, str] = {}
        for arr_key in ("normalized", "redirects"):
            for item in data.get("query", {}).get(arr_key, []) or []:
                src = item.get("from")
                dst = item.get("to")
                if src and dst:
                    alias[src] = dst

        pages = {p.get("title"): p for p in data.get("query", {}).get("pages", [])}

        def resolve_alias(t: str) -> str:
            seen = set()
            while t in alias and t not in seen:
                seen.add(t)
                t = alias[t]
            return t

        for original in group:
            canonical = resolve_alias(original)
            page = pages.get(canonical)
            if page is None:
                candidates = [canonical, canonical.replace(" ", "_"), canonical.replace("_", " ")]
                page = next((pages.get(c) for c in candidates if c in pages), None)

            if not page:
                results[original] = None
                errors.append((original, "No page returned for title"))
                continue

            if page.get("missing"):
                results[original] = None
            else:
                pageid = page.get("pageid")
                results[original] = f"M{pageid}" if pageid else None
                if not pageid:
                    errors.append((original, "Page present but pageid missing"))

    return results, errors


def log_errors_to_csv(
    rows: List[Tuple[str, str]],
    filename: str,
    input_col_label: str = URL_COLUMN,
) -> None:
    """Write (item, error) rows to CSV."""
    if not rows:
        return
    with open(filename, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow([input_col_label, "Error"])
        w.writerows(rows)


def mid_to_entity_url(mid: str) -> str:
    """Map an M-ID like 'M12345' to its human-readable entity page URL."""
    if isinstance(mid, str) and mid.startswith("M") and mid[1:].isdigit():
        return ENTITY_PAGE_BASE + mid
    return "NOT FOUND"


def process() -> None:
    # Load data
    df = pd.read_excel(XLSX_PATH, sheet_name=SHEET_NAME)
    if URL_COLUMN not in df.columns:
        raise KeyError(
            f"Column '{URL_COLUMN}' not found in sheet '{SHEET_NAME}'. "
            f"Available: {list(df.columns)}"
        )

    # Extract titles from URLs (preserve row count)
    urls = df[URL_COLUMN].astype(str)
    titles = [extract_title_from_url(u) for u in urls]

    # Prepare session and fetch M-IDs
    session = build_session()
    nonnull_titles = [t for t in titles if t]
    title_to_mid, api_errors = fetch_mids_for_titles(session, nonnull_titles, BATCH_SIZE)

    # Build final M-ID and URL series aligned to rows
    mids: List[str] = []
    mid_urls: List[str] = []
    errors: List[Tuple[str, str]] = []
    api_error_map: Dict[str, str] = {}
    for title, msg in api_errors:
        api_error_map.setdefault(title, msg)

    for i, url in enumerate(urls):
        t = titles[i]
        if not t:
            mids.append("NOT FOUND")
            mid_urls.append("NOT FOUND")
            errors.append((url, "Could not parse a File: title from URL"))
            continue

        mid = title_to_mid.get(t)
        if mid:
            mids.append(mid)
            mid_urls.append(mid_to_entity_url(mid))
        else:
            mids.append("NOT FOUND")
            mid_urls.append("NOT FOUND")
            reason = api_error_map.get(t, "Page missing or lookup failed")
            errors.append((url, reason))

    # Insert/replace the two output columns right after the URL column
    insert_at = df.columns.get_loc(URL_COLUMN) + 1
    for col_name in (MID_COLUMN, MID_URL_COLUMN):
        if col_name in df.columns:
            df.drop(columns=[col_name], inplace=True)
    df.insert(insert_at, MID_COLUMN, mids)
    df.insert(insert_at + 1, MID_URL_COLUMN, mid_urls)

    # Write back *into the same workbook*, replacing only this sheet
    try:
        with pd.ExcelWriter(
            XLSX_PATH, engine="openpyxl", mode="a", if_sheet_exists="replace"
        ) as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)
    except TypeError:
        # Fallback for older pandas without if_sheet_exists
        from openpyxl import load_workbook
        wb = load_workbook(XLSX_PATH)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
            wb.remove(ws)
        wb.create_sheet(SHEET_NAME)
        wb.save(XLSX_PATH)
        with pd.ExcelWriter(XLSX_PATH, engine="openpyxl", mode="a") as writer:
            df.to_excel(writer, sheet_name=SHEET_NAME, index=False)

    print(f"✅ Updated in place: {XLSX_PATH} (sheet: {SHEET_NAME})")

    # Save errors
    log_errors_to_csv(errors, ERRORS_CSV, input_col_label=URL_COLUMN)
    if errors:
        print(f"⚠️ Errors and failed lookups logged to: {ERRORS_CSV}")
    else:
        print("🎉 No errors encountered.")


if __name__ == "__main__":
    # Tip: make sure the Excel file is closed before running this.
    process()
