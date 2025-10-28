"""
Wikimedia Commons File Metadata Downloader & Excel Writer

Overview
========
This script builds a reliable pipeline to collect **per-file metadata** from
Wikimedia Commons and write it into a single Excel workbook. It supports two
modes:

1) **manual-list** — read a list of files from an input sheet you maintain.
2) **category** — harvest files from a Commons category (optionally a 1-based
   inclusive **range**, e.g., items 20–40), write them to an input sheet, and
   process them exactly like the manual list.

For **every file**, the script:
- Fetches JSON via the Commons API (`prop=imageinfo` with
  `iiprop=extmetadata|url|size|sha1|mime|mediatype|timestamp|user`, `redirects=1`,
  `formatversion=2`, language set by `EXTMETA_LANG`).
- Derives the MediaInfo ID (**MID**) from the returned `pageid` (e.g., `M12345`)
  and a human URL to the entity page.
- Saves the **verbatim JSON** response to `downloaded_metadata/` using a
  **Windows-safe** filename derived from `<CommonsFileName>__<MID or NOID>.json`.
  If the same filename is produced again, it is **overwritten**. If paths risk
  being too long, the base name is truncated and a short hash is added.
- Flattens the JSON to dotted columns and writes rows to an **output sheet** in
  **chunks** (batches), widening columns when new keys appear.

Workbook & sheets
=================
Workbook: **`wmc-inputfiles.xlsx`**

- **Manual-list mode**
  - Input sheet:  **`Files-Manual`**
    - Columns:
      - `CommonsFileName` (required; `File:` prefix is added if missing)
      - `SourceCategory`  (optional; carried through to outputs)
  - Output sheet: **`FilesMetadata-Manual`**
    - De-duplication (hard-wired): rows are deduped by
      (`Input_CommonsFileName`, `Computed_MediaID`).

- **Category mode**
  - Harvest/input sheet: **`Files-Category`**
    - Filled by the script. Each row has:
      - `CommonsFileName` (harvested title, e.g., `File:Example.jpg`)
      - `SourceCategory`  (the category the file came from)
    - **Append-safe harvest**: new category runs APPEND to this sheet, not replace.
      Existing rows are **deduped** by (`CommonsFileName`, `SourceCategory`), so
      you can safely harvest multiple categories into the same sheet.
  - Output sheet: **`FilesMetadata-Category`**
    - De-duplication (hard-wired): rows are deduped by
      (`Input_CommonsFileName`, `Computed_MediaID`, `SourceCategory`).
      This removes true duplicates from re-runs of the same category/range while
      still allowing the same file to appear once per different source category.

Chunking & ranges
=================
- **Processing chunks** (both modes): rows are processed in batches
  (`CHUNK_SIZE`). Each batch is appended to the output sheet; if a batch brings
  new JSON fields, the sheet is widened and replaced once, then appends continue.
- **Harvest flushes** (category mode): harvested filenames buffer in memory and
  are written to `Files-Category` every `HARVEST_FLUSH_ROWS`.
- **Category range**: optional 1-based inclusive slice
  (`CATEGORY_RANGE_START`, `CATEGORY_RANGE_END`) limits which items are taken
  from the category’s traversal order.

Idempotency & overwrite rules
=============================
- **Files-Category**: appends with dedupe on (`CommonsFileName`, `SourceCategory`).
- **FilesMetadata-Manual**: appends with dedupe on
  (`Input_CommonsFileName`, `Computed_MediaID`).
- **FilesMetadata-Category**: appends with dedupe on
  (`Input_CommonsFileName`, `Computed_MediaID`, `SourceCategory`).
- **JSON files**: if the generated filename is the same, it is **overwritten**.

Progress & resilience
=====================
- Per-file progress lines like:
  `[123/8120] Fetching File:Example.jpg … done (MID=M123456)`
- Retry/backoff on transient HTTP errors (429/5xx) with a polite **User-Agent**
  (`USER_AGENT`) per Wikimedia API etiquette.
- If a JSON write fails (e.g., I/O), the row is still written; `Local_JSON_File`
  may be empty, and an error message is printed.
- Input read errors (missing workbook/sheet/columns) raise clear exceptions.

Configuration (edit in code)
============================
A configuration block is included **verbatim** near the top of the script. Key items:

- `MODE` — `"manual-list"` or `"category"`, and (for category mode) `CATEGORY_TITLE`
- `XLSX_PATH` — workbook path (default `wmc-inputfiles.xlsx`)
- Sheet names: `Files-Manual`, `Files-Category`,
  `FilesMetadata-Manual`, `FilesMetadata-Category`
- `CATEGORY_PAGE_LIMIT`, `HARVEST_FLUSH_ROWS`, `CHUNK_SIZE`
- Optional `CATEGORY_RANGE_START` / `CATEGORY_RANGE_END`
- API & file settings: `EXTMETA_LANG`, `USER_AGENT`, `DOWNLOAD_DIR`,
  `FULL_PATH_BUDGET` (conservative Windows full-path limit)

Outputs (columns)
=================
Each output row begins with:
- `Input_CommonsFileName`
- `SourceCategory`
- `Requested_API_URL`
- `Local_JSON_File`
- `Computed_MediaID`
- `Computed_MediaID_URL`
- `BatchIndex` (1-based, the processing chunk number)
…followed by **all flattened JSON fields** (e.g., `query.pages.0.title`,
`query.pages.0.imageinfo.0.url`,
`query.pages.0.imageinfo.0.extmetadata.Artist.value`, …).

How to run
==========
1. Install dependencies (`requirements.txt`) in a virtual environment.
2. Ensure `wmc-inputfiles.xlsx` exists and is **closed** in Excel.
3. Edit the config block (MODE, CATEGORY_TITLE, etc.).
4. Run: `python wmc_metadata.py`

Notes
=====
- `SourceCategory` in manual mode is optional and not modified by the script.
- The output schema grows as new JSON keys appear; earlier rows will have blanks
  for fields not present at the time they were written.
- If you enable long paths on Windows, you may increase `FULL_PATH_BUDGET`.

License & contact
=================
Author: ChatGPT, prompted by Olaf Janssen, KB (Koninklijke Bibliotheek)
Latest update: 2025-10-22
License: CC0
"""

from __future__ import annotations
import hashlib
import json
import math
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import pandas as pd
import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# =========================
# YOUR CONFIGURATION PARAMETERS
# =========================
# Comment out the mode you do NOT want to use:
#MODE = "manual-list" # Comment out if MODE = category
MODE = "category"  # In this mode, you must configure the CATEGORY_TITLE below.
CATEGORY_TITLE = "Category:Media_contributed_by_Koninklijke_Bibliotheek"
#CATEGORY_TITLE = "Category:Catchpenny_prints_from_Koninklijke_Bibliotheek"
# Optional range for CATEGORY mode (1-based inclusive). First 10 files in this category:
CATEGORY_RANGE_START: Optional[int] = 1  # Set to None to process all
CATEGORY_RANGE_END:   Optional[int] = 20  # None = process all

## Harvest settings for MODE = "category"
# Harvest behavior (appends instead of replacing by default)
HARVEST_APPEND: bool = True    # if False, the harvest will overwrite any rows already present in the Files-Category sheet
                               # If True, it keeps existing rows in the Files-Category sheet and appends new ones.
# Only if HARVEST_APPEND: bool = True:
HARVEST_DEDUPE: bool = True    # True = skip rows already present in the Files-Category sheet, as to avoid duplication (by filename+category)

# HARVEST_FLUSH_ROWS : Category-mode only. How many harvested filenames (from the Commons category) we
# accumulate before writing them into the input sheet WMCFiles-Category.
HARVEST_FLUSH_ROWS = 100

# Adapt the User-Agent for with your own details
USER_AGENT = "Wikimedia Commons File Metadata Downloader - User:OlafJanssen - Contact: olaf.janssen@kb.nl)"

# Input workbook - required in both modes
XLSX_PATH = "wmc-inputfiles.xlsx"

# Input sheets
INPUT_SHEET_MANUAL = "Files-Manual"
INPUT_SHEET_CATEGORY = "Files-Category"
# Output sheets
OUTPUT_SHEET_MANUAL = "FilesMetadata-Manual"
OUTPUT_SHEET_CATEGORY = "FilesMetadata-Category"

# Category harvesting
CATEGORY_PAGE_LIMIT = 500  # API max for non-bot

# CHUNK_SIZE → Both modes. How many files we actually process per batch (fetch JSON, save per-file JSON, flatten,
# and append rows to the output sheet).
CHUNK_SIZE = 100  # per your spec

# Where to drop per-file JSON
DOWNLOAD_DIR = Path("downloaded_metadata")

# API & etiquette
COMMONS_API = "https://commons.wikimedia.org/w/api.php"
EXTMETA_LANG = "en"  # or "nl"

# Requests
TIMEOUT_SECS = 20
RETRIES_TOTAL = 5
RETRIES_BACKOFF = 0.6

# Windows path safety
FULL_PATH_BUDGET = 240  # conservative full-path length budget

# =========================

# ---------- HTTP session ----------

def build_session() -> requests.Session:
    """
    Create a requests.Session configured with retry/backoff and a polite User-Agent.

    Returns:
        requests.Session: session preconfigured for Commons requests.

    Notes:
        - Retries on 429/5xx with exponential backoff.
        - UA contains contact info per Wikimedia API etiquette.
    """
    s = requests.Session()
    retry = Retry(
        total=RETRIES_TOTAL,
        connect=RETRIES_TOTAL,
        read=RETRIES_TOTAL,
        status=RETRIES_TOTAL,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset({"GET"}),
        backoff_factor=RETRIES_BACKOFF,
        raise_on_status=False,
    )
    s.mount("https://", HTTPAdapter(max_retries=retry))
    s.headers.update({"User-Agent": USER_AGENT, "Accept": "application/json"})
    return s


# ---------- Utilities ----------

def norm_file_title(name: str) -> str:
    """
    Ensure a Commons title is prefixed with 'File:'.

    Args:
        name: raw title or filename.

    Returns:
        str: normalized title or empty string if input was empty.
    """
    name = (name or "").strip()
    if not name:
        return ""
    return name if name.lower().startswith("file:") else f"File:{name}"


def compute_mid(pageid: Optional[str]) -> str:
    """
    Convert a MediaWiki pageid to a MediaInfo ID (MID).

    Args:
        pageid: page id from the API.

    Returns:
        str: 'M{pageid}' or '' if not numeric/empty.
    """
    return f"M{pageid}" if pageid and str(pageid).isdigit() else ""


def mid_url(mid: str) -> str:
    """
    Build a human-readable MediaInfo entity URL.

    Args:
        mid: MediaInfo id (e.g., 'M12345').

    Returns:
        str: URL or empty string if mid is empty.
    """
    return f"https://commons.wikimedia.org/wiki/Special:EntityPage/{mid}" if mid else ""


def flatten_json(obj: Any, parent_key: str = "", sep: str = ".") -> Dict[str, Any]:
    """
    Flatten a nested dict/list/scalar JSON-like object to a single dict with dotted keys.

    Args:
        obj: dict/list/scalar to flatten.
        parent_key: prefix for recursive calls.
        sep: separator for keys.

    Returns:
        Dict[str, Any]: flattened mapping.

    Safety:
        - Never raises on unexpected shapes; treats unknowns as scalars.
    """
    items: List[Tuple[str, Any]] = []
    if isinstance(obj, dict):
        for k, v in obj.items():
            new_key = f"{parent_key}{sep}{k}" if parent_key else str(k)
            items.extend(flatten_json(v, new_key, sep=sep).items())
    elif isinstance(obj, list):
        for i, v in enumerate(obj):
            new_key = f"{parent_key}{sep}{i}" if parent_key else str(i)
            items.extend(flatten_json(v, new_key, sep=sep).items())
    else:
        items.append((parent_key, obj))
    return dict(items)


def extract_pageid_title(data: Dict[str, Any]) -> Tuple[str, str]:
    """
    Safely read pageid and title from a Commons API 'formatversion=2' response.

    Args:
        data: parsed JSON from API.

    Returns:
        (pageid, title): both '' if missing/missing page.
    """
    try:
        pages = (data.get("query") or {}).get("pages") or []
        if not pages:
            return "", ""
        page = pages[0] or {}
        if page.get("missing"):
            return "", ""
        pid = str(page.get("pageid") or "")
        ttl = str(page.get("title") or "")
        return pid, ttl
    except Exception:
        return "", ""


def safe_component(value: str) -> str:
    """
    Sanitize text for filesystem compatibility (keep alnum, '-', '_', '.').

    Args:
        value: input string.

    Returns:
        str: sanitized or 'NA' if empty.
    """
    value = (value or "").strip()
    if not value:
        return "NA"
    return "".join(c if c.isalnum() or c in ("-", "_", ".") else "_" for c in value)


def short_hash(text: str, n: int = 8) -> str:
    """
    Stable short hash for filename disambiguation.

    Args:
        text: source text to hash.
        n: length of hex digest to return.

    Returns:
        str: first n hex chars of BLAKE2s digest.
    """
    return hashlib.blake2s(text.encode("utf-8"), digest_size=8).hexdigest()[:n]


def build_safe_json_path(
    dir_path: Path,
    input_name: str,
    mid: str,
    budget_full_path: int = FULL_PATH_BUDGET,
    ext: str = ".json",
) -> Path:
    """
    Build a Windows-safe JSON file path derived from input name + MID, keeping
    the FULL PATH below a conservative length budget.

    Behavior:
        - Uses '<input_basename>__<MID or NOID>.json' when possible.
        - If too long, truncates base and adds an 8-char hash.
        - Final fallback: '<MID or NOID>__<hash>.json' or just '<hash>.json'.
        - Deterministic; if the same path is returned twice, the caller will
          OVERWRITE it by opening with mode='w'.

    Args:
        dir_path: directory for the JSON file.
        input_name: original Commons file name (with or without 'File:').
        mid: MediaInfo ID or '' if not found.
        budget_full_path: max allowed length for full path string.
        ext: file extension.

    Returns:
        Path: filesystem path (not created).
    """
    base_raw = (input_name or "").strip()
    base_core = base_raw[5:] if base_raw.lower().startswith("file:") else base_raw
    base = safe_component(base_core) or "NA"
    mid_tag = safe_component(mid or "NOID")

    candidate = f"{base}__{mid_tag}{ext}"

    def within_budget(fname: str) -> bool:
        try:
            return len(str((dir_path / fname).resolve())) <= budget_full_path
        except Exception:
            # If resolve() fails on some FS, fall back to len(dirname)+len(fname)
            return len(str(dir_path)) + 1 + len(fname) <= budget_full_path

    # 1) Try full base
    if within_budget(candidate):
        return dir_path / candidate

    # 2) Add short hash and truncate base to fit
    h = short_hash(base_core)
    suffix = f"__{h}__{mid_tag}{ext}"

    try:
        dir_abs = str(dir_path.resolve())
    except Exception:
        dir_abs = str(dir_path)
    avail_for_fname = max(16, budget_full_path - len(dir_abs) - 1)  # minus path sep
    room_for_base = max(0, avail_for_fname - len(suffix))

    if room_for_base > 0:
        base_trunc = base[:room_for_base]
        candidate2 = f"{base_trunc}{suffix}"
        if within_budget(candidate2):
            return dir_path / candidate2

    # 3) Final fallbacks
    fallback = f"{mid_tag}__{h}{ext}"
    if within_budget(fallback):
        return dir_path / fallback

    tiny = f"{h}{ext}"
    return dir_path / tiny


# ---------- Excel helpers ----------

def sheet_exists(xlsx_path: str, sheet_name: str) -> bool:
    """
    Check if a sheet exists in the workbook.

    Args:
        xlsx_path: path to workbook.
        sheet_name: sheet to check.

    Returns:
        bool: True if sheet is present, False otherwise.
    """
    try:
        with pd.ExcelFile(xlsx_path) as xf:
            return sheet_name in xf.sheet_names
    except FileNotFoundError:
        return False
    except Exception as e:
        print(f"⚠️  Could not read workbook '{xlsx_path}': {e}")
        return False


def read_sheet_df(xlsx_path: str, sheet_name: str) -> pd.DataFrame:
    """
    Read a sheet as DataFrame with dtype=object.

    Raises:
        FileNotFoundError: if workbook is missing.
        ValueError: if sheet missing.
    """
    return pd.read_excel(xlsx_path, sheet_name=sheet_name, dtype="object")


def write_new_sheet(xlsx_path: str, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Create a new sheet (or new workbook if needed) with the given DataFrame.

    Notes:
        - Opens workbook in append mode if file exists.
    """
    try:
        mode = "a" if Path(xlsx_path).exists() else None
        if mode:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as w:
                df.to_excel(w, sheet_name=sheet_name, index=False)
        else:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
                df.to_excel(w, sheet_name=sheet_name, index=False)
    except Exception as e:
        print(f"❌ Failed to write new sheet '{sheet_name}': {e}")
        raise


def replace_sheet(xlsx_path: str, sheet_name: str, df: pd.DataFrame) -> None:
    """
    Replace (or create) a sheet with the provided DataFrame while keeping other sheets intact.

    Compatibility:
        - Uses if_sheet_exists='replace' when available.
        - Falls back to openpyxl removal for older pandas.
    """
    try:
        with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            df.to_excel(w, sheet_name=sheet_name, index=False)
    except TypeError:
        try:
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path)
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                wb.remove(ws)
                wb.save(xlsx_path)
            with pd.ExcelWriter(xlsx_path, engine="openpyxl", mode="a") as w:
                df.to_excel(w, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"❌ Failed to replace sheet '{sheet_name}': {e}")
            raise
    except FileNotFoundError:
        # Workbook doesn't exist yet; create it
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
                df.to_excel(w, sheet_name=sheet_name, index=False)
        except Exception as e:
            print(f"❌ Failed to create workbook '{xlsx_path}': {e}")
            raise
    except Exception as e:
        print(f"❌ Unexpected error replacing sheet '{sheet_name}': {e}")
        raise


def append_chunk_to_sheet(
    xlsx_path: str,
    sheet_name: str,
    chunk: pd.DataFrame,
    dedupe_keys: Optional[List[str]] = None,
    dedupe_keep: str = "first",
) -> None:
    """
    Append a chunk to an existing sheet, widening columns if needed.
    If the sheet doesn't exist, create it with this chunk.

    If dedupe_keys is provided, drop duplicate rows by those keys after
    concatenating existing+chunk. 'dedupe_keep' controls which duplicate
    to retain: 'first' (default) or 'last'.

    Robustness: if some dedupe columns are missing (older runs), we de-dup
    on the subset that exists; if none exist, we skip de-dup with a warning.
    """
    try:
        if not sheet_exists(xlsx_path, sheet_name):
            write_new_sheet(xlsx_path, sheet_name, chunk)
            return

        existing = read_sheet_df(xlsx_path, sheet_name)
        all_cols = list(dict.fromkeys(list(existing.columns) + list(chunk.columns)))
        existing = existing.reindex(columns=all_cols)
        chunk = chunk.reindex(columns=all_cols)

        combined = pd.concat([existing, chunk], ignore_index=True)

        if dedupe_keys:
            present = [k for k in dedupe_keys if k in combined.columns]
            if present:
                combined = combined.drop_duplicates(subset=present, keep=dedupe_keep, ignore_index=True)
            else:
                print(f"⚠️  Dedupe skipped for '{sheet_name}': none of {dedupe_keys} present.")

        replace_sheet(xlsx_path, sheet_name, combined)
    except Exception as e:
        print(f"❌ Failed to append chunk to '{sheet_name}': {e}")
        raise




# ---------- API calls ----------

def commons_params(file_title: str) -> Dict[str, str]:
    """
    Build query parameters for a Commons file metadata request.

    Args:
        file_title: normalized 'File:...' title.

    Returns:
        dict: query parameters for action=query request.
    """
    return {
        "action": "query",
        "format": "json",
        "formatversion": "2",
        "titles": file_title,
        "redirects": "1",
        "prop": "imageinfo",
        "iiprop": "extmetadata|url|size|sha1|mime|mediatype|timestamp|user",
        "iiextmetadatalanguage": EXTMETA_LANG,
    }


def fetch_one(session: requests.Session, file_title: str) -> Tuple[Dict[str, Any], str]:
    """
    Fetch Commons metadata JSON for a single file.

    Args:
        session: configured HTTP session.
        file_title: normalized 'File:...' title.

    Returns:
        (data, effective_url): parsed JSON dict and the fully prepared request URL.

    Raises:
        requests.RequestException: for network/HTTP errors (with context).
    """
    try:
        req = requests.Request("GET", COMMONS_API, params=commons_params(file_title))
        prepped = session.prepare_request(req)
        resp = session.send(prepped, timeout=TIMEOUT_SECS)
        resp.raise_for_status()
        return resp.json(), (prepped.url or "")
    except requests.RequestException as e:
        # Attach title for context
        e.args = (f"{e.args[0] if e.args else e} [title={file_title}]",)
        raise


# ---------- Category harvest (supports optional RANGE) ----------

def harvest_category_to_sheet(
    session: requests.Session,
    category_title: str,
    xlsx_path: str,
    sheet_name: str,
    flush_rows: int = HARVEST_FLUSH_ROWS,
    index_start: Optional[int] = None,  # 1-based inclusive
    index_end: Optional[int] = None,    # 1-based inclusive
    replace_existing: bool = not HARVEST_APPEND,
    dedupe: bool = HARVEST_DEDUPE,
) -> None:
    """
    Harvest files from a Commons category (with continuation) and write to an input sheet.

    Behavior:
      - If replace_existing=True: start with a clean sheet header (REPLACE).
      - Else (default): APPEND to the existing sheet. When dedupe=True, existing (filename, category)
        pairs are skipped so you can safely run multiple categories into the same sheet.
      - Supports selecting a 1-based inclusive RANGE (index_start..index_end). The function
        never clears the sheet for empty ranges; it just warns and returns.

    Columns written: 'CommonsFileName', 'SourceCategory'
    """

    def norm_key(filename: str, source_cat: str) -> tuple[str, str]:
        # Normalize filename to 'File:...' and compare case-insensitively
        fn = norm_file_title((filename or "").strip()).lower()
        sc = (source_cat or "").strip().lower()
        return (fn, sc)

    # Validate range early without touching the sheet
    if index_start is not None and index_end is not None and index_end < index_start:
        print(f"⚠️  Empty range ({index_start}..{index_end}); nothing to harvest. Sheet left unchanged.")
        return

    want_start = index_start or 1
    want_end = index_end or float("inf")

    existing_keys: set[tuple[str, str]] = set()
    total_existing = 0

    if replace_existing:
        # Start fresh: create/replace with just the header row
        replace_sheet(xlsx_path, sheet_name, pd.DataFrame(columns=["CommonsFileName", "SourceCategory"]))
    else:
        # APPEND mode: load existing keys (if any) to support de-duplication
        if sheet_exists(xlsx_path, sheet_name):
            try:
                df_existing = read_sheet_df(xlsx_path, sheet_name)
                # Normalize the expected columns
                if "CommonsFileName" not in df_existing.columns:
                    df_existing["CommonsFileName"] = ""
                if "SourceCategory" not in df_existing.columns:
                    df_existing["SourceCategory"] = ""
                for _, row in df_existing.iterrows():
                    existing_keys.add(norm_key(str(row["CommonsFileName"]), str(row["SourceCategory"])))
                total_existing = len(existing_keys)
            except Exception as e:
                print(f"⚠️  Could not read existing '{sheet_name}' for de-duplication: {e}")

    print(
        f"Harvesting category: {category_title}"
        + (f" [range {index_start}..{index_end}]" if index_start or index_end else "")
        + (" (APPEND mode)" if not replace_existing else " (REPLACE mode)")
        + (", de-dup ON" if dedupe else ", de-dup OFF")
    )

    params = {
        "action": "query",
        "format": "json",
        "list": "categorymembers",
        "cmtitle": category_title,
        "cmtype": "file",
        "cmprop": "title",
        "cmlimit": str(CATEGORY_PAGE_LIMIT),
    }

    harvested_rows: list[dict[str, str]] = []
    total_seen = 0        # files scanned in the category this run
    written_this_run = 0  # appended rows written in this run
    cont: Optional[dict[str, str]] = None

    try:
        while True:
            if cont:
                params.update(cont)
            resp = session.get(COMMONS_API, params=params, timeout=TIMEOUT_SECS)
            resp.raise_for_status()
            data = resp.json()
            members = (data.get("query") or {}).get("categorymembers") or []
            if not members:
                break

            page_first = total_seen + 1
            page_last = total_seen + len(members)

            # Determine overlap with requested range
            take_from = max(want_start, page_first)
            take_to = min(want_end, page_last)

            if take_from <= take_to:
                s = take_from - page_first
                e = take_to - page_first + 1  # inclusive -> exclusive
                subset = members[s:e]
                for m in subset:
                    title = str(m.get("title") or "")
                    if not title:
                        continue
                    row = {"CommonsFileName": title, "SourceCategory": category_title}
                    if dedupe:
                        k = norm_key(row["CommonsFileName"], row["SourceCategory"])
                        if (k in existing_keys):
                            continue
                        existing_keys.add(k)  # reserve now to avoid duplicates within this run
                    harvested_rows.append(row)
                    written_this_run += 1

            total_seen = page_last

            # Flush buffered rows to Excel periodically
            if harvested_rows and (len(harvested_rows) >= flush_rows):
                df_flush = pd.DataFrame(harvested_rows, columns=["CommonsFileName", "SourceCategory"])
                append_chunk_to_sheet(xlsx_path, sheet_name, df_flush)
                harvested_rows.clear()
                print(
                    f"  • Appended {written_this_run} new row(s)"
                    + (f" (existing before run: {total_existing})" if total_existing else "")
                    + f"; scanned {total_seen} items so far…"
                )

            # Stop early when the requested range is complete
            if want_end != float("inf") and written_this_run >= (want_end - want_start + 1):
                break

            cont = data.get("continue")
            if not cont:
                break

    except requests.RequestException as e:
        print(f"❌ Category harvest failed: {e} — partial results kept.")
    except Exception as e:
        print(f"❌ Unexpected error during harvest: {e} — partial results kept.")

    # Final flush
    if harvested_rows:
        try:
            df_flush = pd.DataFrame(harvested_rows, columns=["CommonsFileName", "SourceCategory"])
            append_chunk_to_sheet(xlsx_path, sheet_name, df_flush)
            harvested_rows.clear()
        except Exception as e:
            print(f"❌ Failed to finalize harvest writes: {e}")

    print(
        f"✅ Harvest complete for '{category_title}'. "
        f"Appended {written_this_run} new row(s)"
        + (f"; existing before run: {total_existing}" if total_existing else "")
        + f". Scanned {total_seen} items in category."
    )



# ---------- Processing (chunked) ----------

def process_input_sheet_chunked(
    session: requests.Session,
    xlsx_path: str,
    input_sheet: str,
    output_sheet: str,
    chunk_size: int,
    output_dedupe_keys: Optional[List[str]] = None,   # NEW
    output_dedupe_keep: str = "first",                # NEW
) -> None:
    """
    Process rows from an input sheet in chunks and append results to an output sheet.

    Steps per file:
        - Normalize title to 'File:'.
        - Fetch Commons JSON (with redirects).
        - Derive MID and MID URL.
        - Save full JSON to disk (safe name, overwrite if identical).
        - Flatten JSON and build an output row with base columns + BatchIndex.
        - Append each chunk to the output sheet (widen columns as needed).

    Args:
        session: HTTP session.
        xlsx_path: workbook path.
        input_sheet: the sheet containing 'CommonsFileName' and optional 'SourceCategory'.
        output_sheet: the destination sheet for flattened metadata.
        chunk_size: number of rows per batch.

    Raises:
        FileNotFoundError / ValueError for input read failures.
        Other write errors will be logged and re-raised by helpers.
    """
    try:
        df_in = pd.read_excel(xlsx_path, sheet_name=input_sheet, dtype="string")
    except FileNotFoundError:
        raise FileNotFoundError(f"Input workbook not found: {xlsx_path}")
    except ValueError as e:
        raise ValueError(f"Input sheet '{input_sheet}' not found in {xlsx_path}: {e}")

    if "CommonsFileName" not in df_in.columns:
        raise KeyError(f"Input sheet '{input_sheet}' must contain 'CommonsFileName'.")
    if "SourceCategory" not in df_in.columns:
        df_in["SourceCategory"] = ""

    df_in["CommonsFileName"] = df_in["CommonsFileName"].fillna("").astype(str).str.strip()
    df_in["SourceCategory"] = df_in["SourceCategory"].fillna("").astype(str).str.strip()

    total = len(df_in)
    if total == 0:
        print(f"Nothing to process in '{input_sheet}'.")
        return

    # Ensure JSON dir exists
    try:
        DOWNLOAD_DIR.mkdir(parents=True, exist_ok=True)
    except Exception as e:
        print(f"❌ Cannot create JSON output directory '{DOWNLOAD_DIR}': {e}")
        raise

    num_batches = math.ceil(total / chunk_size)
    print(f"Processing {total} rows from '{input_sheet}' in {num_batches} batch(es) of {chunk_size}…")

    processed_so_far = 0
    for batch_index in range(num_batches):
        start = batch_index * chunk_size
        end = min(start + chunk_size, total)
        batch = df_in.iloc[start:end].copy()

        rows_out: List[Dict[str, Any]] = []
        for i, row in batch.iterrows():
            input_name = (row.get("CommonsFileName") or "").strip()
            source_cat = (row.get("SourceCategory") or "").strip()
            file_title = norm_file_title(input_name)

            print(f"[{i + 1}/{total}] Fetching {file_title or '<EMPTY>'} … ", end="", flush=True)

            if not file_title:
                rows_out.append({
                    "Input_CommonsFileName": input_name,
                    "SourceCategory": source_cat,
                    "Requested_API_URL": "",
                    "Local_JSON_File": "",
                    "Computed_MediaID": "",
                    "Computed_MediaID_URL": "",
                    "BatchIndex": batch_index + 1,
                })
                print("skipped (empty filename).")
                continue

            try:
                data, req_url = fetch_one(session, file_title)
            except requests.RequestException as e:
                rows_out.append({
                    "Input_CommonsFileName": input_name,
                    "SourceCategory": source_cat,
                    "Requested_API_URL": "",
                    "Local_JSON_File": "",
                    "Computed_MediaID": "",
                    "Computed_MediaID_URL": "",
                    "BatchIndex": batch_index + 1,
                    "error.message": str(e),
                })
                print(f"error: {e}")
                continue

            pageid, api_title = extract_pageid_title(data)
            mid = compute_mid(pageid)
            midlink = mid_url(mid)

            # Build JSON path and write (overwrite if same filename)
            try:
                json_path = build_safe_json_path(DOWNLOAD_DIR, input_name or api_title or file_title, mid)
                with json_path.open("w", encoding="utf-8") as f:
                    json.dump(data, f, ensure_ascii=False, indent=4, sort_keys=True)
            except Exception as e:
                # Keep going; record the error in output row
                print(f"⚠️  JSON write failed: {e}")
                json_path = Path("")

            flat = flatten_json(data)
            row_dict: Dict[str, Any] = {
                "Input_CommonsFileName": input_name,
                "SourceCategory": source_cat,
                "Requested_API_URL": req_url,
                "Local_JSON_File": str(json_path),
                "Computed_MediaID": mid,
                "Computed_MediaID_URL": midlink,
                "BatchIndex": batch_index + 1,
            }
            row_dict.update(flat)
            rows_out.append(row_dict)

            print(f"done ({'MID=' + mid if mid else 'MID=NOT FOUND'}).")

        # Build chunk DF with base columns first
        chunk_df = pd.DataFrame(rows_out)
        front_cols = [
            "Input_CommonsFileName",
            "SourceCategory",
            "Requested_API_URL",
            "Local_JSON_File",
            "Computed_MediaID",
            "Computed_MediaID_URL",
            "BatchIndex",
        ]
        cols = [c for c in front_cols if c in chunk_df.columns] + [c for c in chunk_df.columns if c not in front_cols]
        chunk_df = chunk_df.reindex(columns=cols)

        # Append/widen/replace output sheet
        append_chunk_to_sheet(
            xlsx_path,
            output_sheet,
            chunk_df,
            dedupe_keys=output_dedupe_keys,
            dedupe_keep=output_dedupe_keep,
        )

        processed_so_far += len(batch)
        print(f"[Batch {batch_index + 1}/{num_batches}] Wrote {len(batch)} rows → '{output_sheet}' (total {processed_so_far}/{total}).")

    print(f"✅ Processing complete → '{output_sheet}'.")


# ---------- Orchestration ----------

def run_manual_list() -> None:
    """
    Execute the manual-list mode: read 'Files-Manual' and write metadata to 'FilesMetadata-Manual'.
    De-duplicate output rows by (Input_CommonsFileName, Computed_MediaID).
    """
    session = build_session()
    process_input_sheet_chunked(
        session=session,
        xlsx_path=XLSX_PATH,
        input_sheet=INPUT_SHEET_MANUAL,
        output_sheet=OUTPUT_SHEET_MANUAL,
        chunk_size=CHUNK_SIZE,
        output_dedupe_keys=["Input_CommonsFileName", "Computed_MediaID"],  # hard-wired, these are the column names in the output sheet.)
        output_dedupe_keep="first",  # keep the existing row if duplicate appears
    )


def run_category() -> None:
    session = build_session()

    # Harvest (your existing call, unchanged)
    harvest_category_to_sheet(
        session=session,
        category_title=CATEGORY_TITLE,
        xlsx_path=XLSX_PATH,
        sheet_name=INPUT_SHEET_CATEGORY,
        flush_rows=HARVEST_FLUSH_ROWS,
        index_start=CATEGORY_RANGE_START,
        index_end=CATEGORY_RANGE_END,
        # keep your existing append/dedupe flags for the *input* sheet
    )

    # Process → de-dup output by (filename, MID, source category)
    process_input_sheet_chunked(
        session=session,
        xlsx_path=XLSX_PATH,
        input_sheet=INPUT_SHEET_CATEGORY,
        output_sheet=OUTPUT_SHEET_CATEGORY,
        chunk_size=CHUNK_SIZE,
        output_dedupe_keys=["Input_CommonsFileName", "Computed_MediaID", "SourceCategory"], # hard-wired, these are the column names in the output sheet.)
        output_dedupe_keep="first",  # or "last" if you prefer latest to win
    )


def main() -> None:
    """
    Entry point. Reads MODE and runs the corresponding orchestration.
    """
    if MODE == "manual-list":
        print("Mode: manual-list")
        run_manual_list()
    elif MODE == "category":
        print("Mode: category")
        run_category()
    else:
        raise ValueError(f"Unknown MODE: {MODE!r}. Use 'manual-list' or 'category'.")


if __name__ == "__main__":
    # Ensure the workbook is closed before running.
    main()
